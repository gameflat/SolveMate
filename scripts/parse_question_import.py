import csv
import json
import pathlib
import re
import sys
import zipfile
import xml.etree.ElementTree as ET

import openpyxl


OPTION_KEYS = list("ABCDEFG")
TYPE_LABELS = {
    "single": "单选题",
    "multiple": "多选题",
    "judge": "判断题",
    "fill": "填空题",
    "short": "简答题",
    "unknown": "其他",
}
TYPE_KEYWORDS = {
    "single": ["单选", "单项", "选择题"],
    "multiple": ["多选", "多项"],
    "judge": ["判断", "对错", "是非"],
    "fill": ["填空"],
    "short": ["简答", "问答", "计算", "论述", "分析"],
}
HEADER_ALIASES = {
    "prompt": ["题干", "题目", "问题", "内容", "question", "prompt"],
    "type": ["题型", "类型", "type"],
    "answer": ["答案", "标准答案", "正确答案", "answer"],
}


def clean(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\u200b", "").replace("\u200c", "")).strip()


def normalize_text(value):
    return str(value or "").replace("\u200b", "").replace("\u200c", "").replace("\r\n", "\n").replace("\r", "\n").strip()


def answer_keys(answer, question_type):
    if question_type in ("single", "multiple"):
        return list(re.sub("[^A-Ga-g]", "", answer).upper())
    return []


def normalize_type(raw_type, prompt="", answer="", options=None):
    text = f"{raw_type} {prompt}".lower()
    for question_type, keywords in TYPE_KEYWORDS.items():
        if any(keyword.lower() in text for keyword in keywords):
            return question_type
    options = options or []
    if options:
        keys = answer_keys(answer, "single")
        return "multiple" if len(keys) > 1 else "single"
    if re.search(r"正确|错误|对|错|√|×|是|否", answer) and re.search(r"[（(]\s*[）)]|判断|正确|错误", prompt):
        return "judge"
    if re.search(r"_{2,}|填空|空[一二三四五六七八九十\d]", prompt):
        return "fill"
    return "short" if prompt else "unknown"


def raw_type_for(question_type, raw_type=""):
    return clean(raw_type) or TYPE_LABELS.get(question_type, "其他")


def normalize_judge_answer(answer):
    value = clean(answer)
    if re.search(r"√|对|正确|是|true", value, re.I):
        return "正确"
    if re.search(r"×|错|错误|否|false", value, re.I):
        return "错误"
    return value


def strip_number(text):
    return re.sub(r"^\s*(第?\d+\s*[题、.．)]|\d+\s*[、.．)]|[一二三四五六七八九十]+[、.．)])\s*", "", clean(text))


def find_header_map(row):
    normalized = [clean(cell).lower() for cell in row]
    mapping = {}
    for target, aliases in HEADER_ALIASES.items():
        for index, cell in enumerate(normalized):
            if any(alias.lower() in cell for alias in aliases):
                mapping[target] = index
                break
    for key in OPTION_KEYS:
        for index, cell in enumerate(normalized):
            compact = re.sub(r"\s+", "", cell).upper()
            if compact in (key, f"选项{key}", f"{key}选项", f"OPTION{key}"):
                mapping[key] = index
                break
    return mapping if "prompt" in mapping else {}


def row_value(row, index):
    if index is None or index >= len(row):
        return ""
    return clean(row[index])


def parse_table_rows(rows, source_name):
    warnings = []
    questions = []
    header = {}
    start_index = 0

    for index, row in enumerate(rows[:10]):
        candidate = find_header_map(row)
        if candidate:
            header = candidate
            start_index = index + 1
            break

    if not header:
        header = {"prompt": 0, "type": 1, "answer": 9}
        for offset, key in enumerate(OPTION_KEYS, start=2):
            header[key] = offset

    for row_number, row in enumerate(rows[start_index:], start=start_index + 1):
        prompt = row_value(row, header.get("prompt"))
        if not prompt:
            continue
        raw_type = row_value(row, header.get("type"))
        answer = row_value(row, header.get("answer"))
        options = []
        for key in OPTION_KEYS:
            text = row_value(row, header.get(key))
            if text:
                options.append({"key": key, "text": text})

        if not answer:
            for cell in reversed(row):
                value = clean(cell)
                if value and value != prompt and value != raw_type and value not in [item["text"] for item in options]:
                    answer = value
                    break

        question_type = normalize_type(raw_type, prompt, answer, options)
        if question_type == "judge":
            options = [{"key": "A", "text": "正确"}, {"key": "B", "text": "错误"}]
            answer = normalize_judge_answer(answer)

        if not answer:
            warnings.append({"row": row_number, "reason": "missing answer", "prompt": prompt[:120]})
        if question_type in ("single", "multiple") and len(options) < 2:
            warnings.append({"row": row_number, "reason": "choice question has fewer than two options", "prompt": prompt[:120]})

        questions.append(
            {
                "sourceIndex": len(questions) + 1,
                "excelRow": row_number,
                "prompt": strip_number(prompt),
                "rawType": raw_type_for(question_type, raw_type),
                "type": question_type,
                "options": options,
                "answer": answer,
                "answerKeys": answer_keys(answer, question_type),
                "source": source_name,
            }
        )

    return questions, warnings


def load_excel(path):
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    rows = []
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        for row in worksheet.iter_rows(values_only=True):
            rows.append([cell if cell is not None else "" for cell in row])
    return rows


def load_csv(path):
    text = pathlib.Path(path).read_text(encoding="utf-8-sig", errors="ignore")
    return list(csv.reader(text.splitlines()))


def load_docx_text(path):
    with zipfile.ZipFile(path) as archive:
        parts = ["word/document.xml"]
        parts.extend(name for name in archive.namelist() if name.startswith("word/tables/table") and name.endswith(".xml"))
        lines = []
        namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
        for part in parts:
            if part not in archive.namelist():
                continue
            root = ET.fromstring(archive.read(part))
            for paragraph in root.findall(".//w:p", namespace):
                texts = [node.text or "" for node in paragraph.findall(".//w:t", namespace)]
                line = normalize_text("".join(texts))
                if line:
                    lines.append(line)
        return "\n".join(lines)


def load_plain_text(path):
    raw = pathlib.Path(path).read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="ignore")


def split_answer(line):
    patterns = [
        r"【\s*(?:参考)?答案\s*】\s*(.+)$",
        r"(?:参考答案|标准答案|正确答案|答案)\s*[:：]\s*(.+)$",
        r"答\s*[:：]\s*(.+)$",
    ]
    for pattern in patterns:
        match = re.search(pattern, line, re.I)
        if match:
            before = line[: match.start()].strip()
            return before, clean(match.group(1))
    return line, None


def is_question_start(line):
    return bool(re.match(r"^\s*(第?\d+\s*[题、.．)]|\d+\s*[、.．)]|[一二三四五六七八九十]+[、.．)])\s*", line))


def split_text_blocks(text):
    lines = [normalize_text(line) for line in text.splitlines()]
    lines = [line for line in lines if line]
    blocks = []
    current = []
    has_answer = False
    for line in lines:
        _, answer = split_answer(line)
        starts_new = is_question_start(line) and current and has_answer
        if starts_new:
            blocks.append(current)
            current = []
            has_answer = False
        current.append(line)
        if answer is not None:
            has_answer = True
    if current:
        blocks.append(current)

    if len(blocks) <= 1:
        paragraph_blocks = []
        current = []
        for raw_line in text.splitlines():
            line = normalize_text(raw_line)
            if not line:
                if current:
                    paragraph_blocks.append(current)
                    current = []
                continue
            current.append(line)
        if current:
            paragraph_blocks.append(current)
        if len(paragraph_blocks) > len(blocks):
            return paragraph_blocks
    return blocks


def parse_options(text):
    matches = list(re.finditer(r"(?<![A-Za-z0-9])([A-G])\s*[.．、)]\s*", text))
    if len(matches) < 2:
        lines = [line for line in text.splitlines() if line.strip()]
        options = []
        prompt_lines = []
        for line in lines:
            match = re.match(r"^\s*([A-G])\s*[.．、)]?\s*(.+)$", line)
            if match:
                options.append({"key": match.group(1), "text": clean(match.group(2))})
            else:
                prompt_lines.append(line)
        return strip_number(" ".join(prompt_lines)), options

    prompt = text[: matches[0].start()]
    options = []
    for index, match in enumerate(matches):
        start = match.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        option_text = clean(text[start:end])
        if option_text:
            options.append({"key": match.group(1).upper(), "text": option_text})
    return strip_number(prompt), dedupe_options(options)


def dedupe_options(options):
    seen = set()
    result = []
    for option in options:
        if option["key"] in seen:
            continue
        seen.add(option["key"])
        result.append(option)
    return result


def parse_text_questions(text, source_name):
    warnings = []
    questions = []
    for block_index, block in enumerate(split_text_blocks(text), start=1):
        raw_lines = []
        answer = ""
        raw_type = ""
        for line in block:
            before, found_answer = split_answer(line)
            if found_answer is not None:
                if before:
                    raw_lines.append(before)
                answer = found_answer
                continue
            if re.search(r"题型|类型", line) and re.search(r"单选|多选|判断|填空|简答|问答", line):
                raw_type = line
                continue
            raw_lines.append(line)

        question_text = "\n".join(raw_lines).strip()
        prompt, options = parse_options(question_text)
        question_type = normalize_type(raw_type, prompt or question_text, answer, options)
        if not prompt:
            prompt = strip_number(question_text)
        if question_type == "judge":
            options = [{"key": "A", "text": "正确"}, {"key": "B", "text": "错误"}]
            answer = normalize_judge_answer(answer)
        if not prompt or not answer:
            warnings.append({"block": block_index, "reason": "missing prompt or answer", "text": " ".join(block)[:180]})
        if prompt and answer:
            questions.append(
                {
                    "sourceIndex": len(questions) + 1,
                    "prompt": prompt,
                    "rawType": raw_type_for(question_type, raw_type),
                    "type": question_type,
                    "options": options,
                    "answer": answer,
                    "answerKeys": answer_keys(answer, question_type),
                    "source": source_name,
                }
            )
    return questions, warnings


def parse_file(path):
    suffix = pathlib.Path(path).suffix.lower()
    source_name = pathlib.Path(path).name
    warnings = []
    source_text = ""

    if suffix in (".xlsx", ".xlsm"):
        rows = load_excel(path)
        questions, row_warnings = parse_table_rows(rows, source_name)
        source_text = "\n".join("\t".join(clean(cell) for cell in row) for row in rows[:400])
        warnings.extend(row_warnings)
    elif suffix == ".csv":
        rows = load_csv(path)
        questions, row_warnings = parse_table_rows(rows, source_name)
        source_text = "\n".join(",".join(clean(cell) for cell in row) for row in rows[:400])
        warnings.extend(row_warnings)
    elif suffix == ".docx":
        source_text = load_docx_text(path)
        questions, text_warnings = parse_text_questions(source_text, source_name)
        warnings.extend(text_warnings)
    elif suffix in (".txt", ".md"):
        source_text = load_plain_text(path)
        questions, text_warnings = parse_text_questions(source_text, source_name)
        warnings.extend(text_warnings)
    else:
        raise ValueError(f"unsupported file type: {suffix or 'unknown'}")

    return {
        "questions": questions,
        "warnings": warnings[:100],
        "sourceText": source_text[:30000],
    }


def main():
    if len(sys.argv) < 2:
        raise SystemExit("usage: parse_question_import.py <file>")
    payload = parse_file(sys.argv[1])
    print(json.dumps(payload, ensure_ascii=False))


if __name__ == "__main__":
    main()

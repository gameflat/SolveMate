import hashlib
import json
import pathlib
import re

import openpyxl
from docx import Document


OLD_BANK_PATH = pathlib.Path("Question Bank/竞赛题库5.22.xlsx")
NEW_BANK_PATH = pathlib.Path("Question Bank/烯烃事业部题库（MTO装置LORU单元）.docx")
BANKS_OUTPUT_PATH = pathlib.Path("data/question-banks.json")
LEGACY_OUTPUT_PATH = pathlib.Path("data/questions.json")
IMPORT_REPORT_PATH = pathlib.Path("data/import-report.json")
OPTION_KEYS = list("ABCDEFG")
TYPE_MAP = {
    "单选题": "single",
    "多选题": "multiple",
    "判断题": "judge",
    "填空题": "fill",
    "简答题": "short",
}
DOCX_SECTION_TYPES = {
    "一、判断题": ("判断题", "judge"),
    "二、单选题": ("单选题", "single"),
    "三、多选题": ("多选题", "multiple"),
    "四、简答题": ("简答题", "short"),
    "五、计算题": ("计算题", "short"),
    "六、论述题": ("论述题", "short"),
}


def clean(value):
    return re.sub(r"\s+", " ", str(value or "").replace("\u200c", "").replace("\u200b", "")).strip()


def normalize_doc_text(value):
    return str(value or "").replace("\u200c", "").replace("\u200b", "").strip()


def stable_id(bank_id, source_index, raw_type, prompt):
    return hashlib.sha1(f"{bank_id}:{source_index}:{raw_type}:{prompt}".encode("utf-8")).hexdigest()[:12]


def answer_keys(answer, question_type):
    if question_type in ("single", "multiple"):
        return list(re.sub("[^A-Ga-g]", "", answer).upper())
    return []


def load_old_excel_bank():
    workbook = openpyxl.load_workbook(OLD_BANK_PATH, data_only=True, read_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    questions = []

    for excel_row, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        prompt = clean(row[0] if len(row) > 0 else "")
        raw_type = clean(row[1] if len(row) > 1 else "")
        if not prompt or not raw_type:
            continue

        question_type = TYPE_MAP.get(raw_type, "unknown")
        options = []
        for offset, key in enumerate(OPTION_KEYS, start=2):
            text = clean(row[offset] if len(row) > offset else "")
            if text:
                options.append({"key": key, "text": text})

        answer = clean(row[9] if len(row) > 9 else "") or clean(row[2] if len(row) > 2 else "")
        question_id = hashlib.sha1(f"{excel_row}:{raw_type}:{prompt}".encode()).hexdigest()[:12]
        questions.append(
            {
                "id": question_id,
                "bankId": "legacy-competition-20240522",
                "excelRow": excel_row,
                "sourceIndex": excel_row,
                "prompt": prompt,
                "rawType": raw_type,
                "type": question_type,
                "options": options,
                "answer": answer,
                "answerKeys": answer_keys(answer, question_type),
            }
        )

    return {
        "id": "legacy-competition-20240522",
        "name": "竞赛题库5.22",
        "label": "过往题库",
        "source": str(OLD_BANK_PATH),
        "isLegacy": True,
        "questions": questions,
    }


def load_docx_bank():
    document = Document(str(NEW_BANK_PATH))
    all_lines = [normalize_doc_text(paragraph.text) for paragraph in document.paragraphs]
    lines = [line for line in all_lines if line]

    first_content_heading = next(
        index
        for index, line in enumerate(lines)
        if line == "一、判断题" and index > 0 and not re.search(r"\t|\\s\\d+$", lines[index])
    )
    lines = lines[first_content_heading:]

    questions = []
    warnings = []
    section_name = ""
    raw_type = ""
    question_type = "unknown"
    question_lines = []
    current = None
    source_index = 0

    for index, line in enumerate(lines):
        if line in DOCX_SECTION_TYPES:
            if current:
                questions.append(finalize_docx_question(current, section_name, warnings))
                current = None
            question_lines = []
            section_name = line
            raw_type, question_type = DOCX_SECTION_TYPES[line]
            continue

        if not section_name:
            continue

        before_answer, answer = split_answer_line(line)
        if answer is not None:
            if before_answer:
                question_lines.append(before_answer)
            source_index += 1
            current = {
                "sourceIndex": source_index,
                "section": section_name,
                "rawType": raw_type,
                "type": question_type,
                "questionLines": question_lines,
                "answerLines": [answer],
            }
            question_lines = []
            continue

        if current:
            next_line = lines[index + 1] if index + 1 < len(lines) else ""
            if starts_next_question(line, next_line, question_type):
                questions.append(finalize_docx_question(current, section_name, warnings))
                current = None
                question_lines = [line]
            else:
                current["answerLines"].append(line)
            continue

        question_lines.append(line)

    if current:
        questions.append(finalize_docx_question(current, section_name, warnings))

    for question in questions:
        question["bankId"] = "mto-loru-202103"
        question["id"] = stable_id("mto-loru-202103", question["sourceIndex"], question["rawType"], question["prompt"])
        question["answerKeys"] = answer_keys(question["answer"], question["type"])

    return {
        "id": "mto-loru-202103",
        "name": "烯烃事业部题库（MTO装置LORU单元）",
        "label": "当前题库",
        "source": str(NEW_BANK_PATH),
        "isLegacy": False,
        "questions": questions,
        "warnings": warnings,
    }


def split_answer_line(line):
    if "【答案" not in line:
        return line, None
    before, marker_and_after = line.split("【答案", 1)
    after = marker_and_after
    if "】" in after:
        after = after.split("】", 1)[1]
    return before.strip(), clean_answer(after)


def clean_answer(value):
    return normalize_doc_text(value).replace("】", "").strip()


def starts_next_question(line, next_line, question_type):
    if split_answer_line(line)[1] is not None:
        return False
    if question_type in ("judge", "single", "multiple"):
        return True
    if re.match(r"^\d+\s*[.．、]", line):
        return True
    if "【答案" in next_line and re.search(r"[?？。)]$", line):
        return True
    return False


def finalize_docx_question(item, section_name, warnings):
    raw_type = item["rawType"]
    question_type = item["type"]
    question_text = "\n".join(part for part in item["questionLines"] if part.strip()).strip()
    answer = "\n".join(part for part in item["answerLines"] if part.strip()).strip()

    if question_type in ("single", "multiple"):
        prompt, options = split_choice_prompt_options(question_text)
    elif question_type == "judge":
        prompt = clean_prompt_number(question_text)
        options = [{"key": "A", "text": "正确"}, {"key": "B", "text": "错误"}]
        answer = normalize_judge_answer(answer)
    else:
        prompt = clean_prompt_number(question_text)
        options = []

    if not prompt or not answer:
        warnings.append(
            {
                "sourceIndex": item["sourceIndex"],
                "section": section_name,
                "reason": "empty prompt or answer",
                "questionText": question_text[:200],
                "answer": answer[:200],
            }
        )
    if question_type in ("single", "multiple") and len(options) < 2:
        warnings.append(
            {
                "sourceIndex": item["sourceIndex"],
                "section": section_name,
                "reason": "choice question has fewer than two parsed options",
                "questionText": question_text[:300],
            }
        )

    return {
        "sourceIndex": item["sourceIndex"],
        "prompt": prompt,
        "rawType": raw_type,
        "type": question_type,
        "options": options,
        "answer": answer,
    }


def normalize_judge_answer(answer):
    if "√" in answer or "对" in answer or "正确" in answer:
        return "正确"
    if "×" in answer or "错" in answer or "错误" in answer:
        return "错误"
    return answer


def clean_prompt_number(text):
    return re.sub(r"^\d+\s*[.．、]\s*", "", clean(text))


def split_choice_prompt_options(text):
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    if len(lines) >= 3 and not any(re.search(r"(?<![A-Za-z0-9])([A-G])\s*[.．、]", line) for line in lines[1:]):
        return fallback_line_options(lines)

    prompt_parts = []
    options = []
    started_options = False

    for line_index, line in enumerate(lines):
        explicit_matches = list(re.finditer(r"(?<![A-Za-z0-9])([A-G])\s*[.．、]\s*", line))
        loose_start = re.match(r"^([A-G])(?=\s|[\u4e00-\u9fff])\s*", line) if line_index > 0 else None
        matches = explicit_matches or ([loose_start] if loose_start and started_options else [])

        if not matches:
            if started_options and options:
                options[-1]["text"] = clean(f"{options[-1]['text']} {line}")
            else:
                prompt_parts.append(line)
            continue

        first = matches[0]
        before = line[: first.start()].strip()
        if before and not started_options:
            prompt_parts.append(before)
        started_options = True

        for index, match in enumerate(matches):
            start = match.end()
            end = matches[index + 1].start() if index + 1 < len(matches) else len(line)
            option_text = clean(line[start:end])
            if option_text:
                options.append({"key": match.group(1).upper(), "text": option_text})

    if len(options) < 2 and len(lines) >= 3:
        return fallback_line_options(lines)

    deduped = []
    seen = set()
    for option in options:
        if option["key"] in seen:
            continue
        seen.add(option["key"])
        deduped.append(option)
    return clean_prompt_number(" ".join(prompt_parts)), deduped


def fallback_line_options(lines):
    prompt = clean_prompt_number(lines[0])
    options = []
    for key, line in zip(OPTION_KEYS, lines[1:]):
        text = re.sub(r"^[A-G]\s*[.．、]?\s*", "", clean(line))
        if text:
            options.append({"key": key, "text": text})
    return prompt, options


def main():
    old_bank = load_old_excel_bank()
    new_bank = load_docx_bank()
    payload = {
        "defaultBankId": new_bank["id"],
        "banks": [new_bank, old_bank],
    }
    BANKS_OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    BANKS_OUTPUT_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    LEGACY_OUTPUT_PATH.write_text(json.dumps(new_bank["questions"], ensure_ascii=False, indent=2), encoding="utf-8")
    IMPORT_REPORT_PATH.write_text(
        json.dumps(
            {
                "banks": [
                    {"id": new_bank["id"], "questions": len(new_bank["questions"]), "warnings": len(new_bank["warnings"])},
                    {"id": old_bank["id"], "questions": len(old_bank["questions"]), "warnings": 0},
                ],
                "warnings": new_bank["warnings"],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    print(f"Imported {len(new_bank['questions'])} new questions and {len(old_bank['questions'])} legacy questions.")
    print(f"Wrote {BANKS_OUTPUT_PATH} and {IMPORT_REPORT_PATH}.")


if __name__ == "__main__":
    main()
